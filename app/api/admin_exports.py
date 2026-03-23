from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from ..config import settings
from ..database import get_db
from ..models import EventLog, ExportRecord, Project, Questionnaire, SurveyResponse, SurveySession, utcnow
from ..security import get_current_admin
from ..services.projects import PROJECT_PURGED
from ..services.questionnaire import build_canonical_order


router = APIRouter(
    prefix="/admin/exports",
    tags=["admin-exports"],
    dependencies=[Depends(get_current_admin)],
)

META_COLUMNS = ["participant_id", "session_id", "status", "started_at", "completed_at"]


def _as_cell_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


@router.get("/{project_id}")
def export_project_data(
    project_id: str,
    questionnaire_id: str | None = Query(default=None),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project or project.delete_status == PROJECT_PURGED:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    questionnaire_query = db.query(Questionnaire).filter(Questionnaire.project_id == project_id)
    if questionnaire_id:
        questionnaire = questionnaire_query.filter(Questionnaire.id == questionnaire_id).first()
    else:
        questionnaire = questionnaire_query.order_by(Questionnaire.created_at.desc()).first()
    if not questionnaire:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Questionnaire not found")

    canonical_order = build_canonical_order(questionnaire.structure_json)
    sessions = (
        db.query(SurveySession)
        .filter(SurveySession.project_id == project_id, SurveySession.questionnaire_id == questionnaire.id)
        .order_by(SurveySession.started_at.asc())
        .all()
    )

    wb = Workbook()
    ws_answer = wb.active
    ws_answer.title = "答题"
    ws_duration = wb.create_sheet("时长")
    ws_order = wb.create_sheet("顺序")

    headers = META_COLUMNS + canonical_order
    ws_answer.append(headers)
    ws_duration.append(headers)
    ws_order.append(headers)

    for session in sessions:
        responses = (
            db.query(SurveyResponse)
            .filter(SurveyResponse.session_id == session.id)
            .order_by(SurveyResponse.present_position.asc())
            .all()
        )
        response_map = {row.item_id: row for row in responses}
        jump_events = (
            db.query(EventLog)
            .filter(EventLog.session_id == session.id, EventLog.event_type == "logic_jump")
            .order_by(EventLog.server_ts.asc())
            .all()
        )
        logic_skipped: set[str] = set()
        for event in jump_events:
            payload = event.payload_json or {}
            skipped = payload.get("skipped_item_ids")
            if isinstance(skipped, list):
                logic_skipped.update(str(item_id) for item_id in skipped)

        meta_values = [
            session.participant_id,
            session.id,
            session.status,
            session.started_at.isoformat() if session.started_at else None,
            session.completed_at.isoformat() if session.completed_at else None,
        ]
        answer_row = list(meta_values)
        duration_row = list(meta_values)
        order_row = list(meta_values)

        for item_id in canonical_order:
            row = response_map.get(item_id)
            if row is None:
                if item_id in logic_skipped:
                    answer_row.append(-1)
                    duration_row.append(-1)
                    order_row.append(-1)
                else:
                    answer_row.append(None)
                    duration_row.append(None)
                    order_row.append(None)
            else:
                answer_row.append(_as_cell_value(row.answer_json))
                duration_row.append(row.duration_ms)
                order_row.append(row.present_position)

        ws_answer.append(answer_row)
        ws_duration.append(duration_row)
        ws_order.append(order_row)

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()

    export_dir = Path(settings.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{project.project_key}_{questionnaire.id}_{timestamp}.xlsx"
    file_path = export_dir / filename
    file_path.write_bytes(content)

    record = ExportRecord(project_id=project_id, file_name=filename, file_path=str(file_path))
    db.add(record)
    db.commit()

    stream = BytesIO(content)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
